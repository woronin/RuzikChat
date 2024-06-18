import datetime
import json
import logging
import os
import random
import redis
import requests
import urllib.request
from dotenv import load_dotenv

import pandas as pd
import openpyxl
import vk_api
from vk_api.bot_longpoll import VkBotLongPoll, VkBotEventType
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
from vk_api.utils import get_random_id

from database.admin_redis_tools import (
    clear_answers,
    clear_qa_from_dir,
    clear_questions,
    delete_select_correct_answer,
    delete_select_question,
    upload_all_files_of_qa,
    upload_one_file_of_qa
)

from database.client_redis_tools import (
    get_last_correct_answer,
    get_last_question,
    get_user_qa,
    get_users_info,
    scale_text,
    upload_account,
    upload_last_correct_answer,
    upload_last_question
)

from giga_chat.giga_model import (
    connect_ruzik_chat,
    custom_evaluate_qa,
    custom_generate_qa,
    get_cosine_similarity,
    get_token
)
from langchain_community.document_loaders import TextLoader
from vk.vk_tools import get_user_info

logger = logging.getLogger('vk bot')

def create_keyboard():
    # одноразовая клавиатура
    keyboard = VkKeyboard(one_time=True)
    keyboard.add_button("Вопрос", color=VkKeyboardColor.PRIMARY)
    keyboard.add_button("Стоп", color=VkKeyboardColor.NEGATIVE)

    keyboard.add_line()
    keyboard.add_button("На счете", color=VkKeyboardColor.POSITIVE)

    return keyboard.get_keyboard()

def create_admin_keyboard():
    # одноразовая клавиатура
    keyboard = VkKeyboard(one_time=True)
    keyboard.add_button("Данные игроков", color=VkKeyboardColor.PRIMARY)
    keyboard.add_button("Вопросы и ответы", color=VkKeyboardColor.NEGATIVE)
    keyboard.add_line()
    keyboard.add_button("Генерация вопросов", color=VkKeyboardColor.POSITIVE)
    keyboard.add_line()
    keyboard.add_button("Изменить логин и пароль", color=VkKeyboardColor.SECONDARY)

    return keyboard.get_keyboard()

def create_qa_admin_keyboard():
    keyboard = VkKeyboard(one_time=True)
    keyboard.add_button("Добавлять вопросы и ответы", color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button("Редактировать вопросы и ответы", color=VkKeyboardColor.NEGATIVE)
    keyboard.add_line()
    keyboard.add_button("Удалять вопросы и ответы", color=VkKeyboardColor.POSITIVE)
    keyboard.add_line()
    keyboard.add_button("Назад", color=VkKeyboardColor.SECONDARY)
    return keyboard.get_keyboard()

def create_request_admin_keyboard():
    keyboard = VkKeyboard(one_time=True)
    keyboard.add_button("Да", color=VkKeyboardColor.POSITIVE)
    keyboard.add_button("Нет", color=VkKeyboardColor.NEGATIVE)
    return keyboard.get_keyboard()

def handle_user_start(vk, user_id):
    message = "Начинаем викторину!"
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_keyboard())
def user_func_start(VK_USER_TOKEN, user_id, r_conn):
    logger.info(f'the user_{user_id} started quiz.')
    # только первый раз составляет список вопросов для каждого пользователя
    if not r_conn.hgetall(f'questions_id_{user_id}'):
        upload_all_files_of_qa(r_conn, directory_name='questions_data', user_id=user_id)

    if not r_conn.hgetall('users_info') or r_conn.hget('users_info', f'id_{user_id}') is None:
        first_name, last_name, b_date, sex, city, country = get_user_info(user_id, VK_USER_TOKEN)
        info_dict = {'first_name': first_name,
                     'last_name': last_name,
                     'b_date': b_date,
                     'sex': sex,
                     'city': city,
                     'country': country,
                     'account': 0
                     }

        r_conn.hset('users_info',
                    f'id_{user_id}',
                    json.dumps(info_dict))
def handle_user_stop(vk, user_id):
    logger.info(f'the user_{user_id} stopped bot.')
    message = "Вышли из виткорины!"
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id())

def choice_question(user_id, r_conn):
    nums_question = r_conn.hkeys(f'questions_id_{user_id}')

    if not nums_question:
        return None

    num_question = random.choice(nums_question)
    return num_question

def func_question(user_id, r_conn):

    num_qa = choice_question(user_id, r_conn)

    if num_qa is not None:
        # сохраняем последний вопрос пользователя
        r_conn.hset(
            'num_of_last_question',
            f'id_{user_id}',
            num_qa
        )

        question = r_conn.hget(f'questions_id_{user_id}', num_qa)
        # answer_num = question_num.replace('question', 'answer')
        answer = r_conn.hget(f'correct_answers_id_{user_id}', num_qa)

        logger.info(f"the user_{user_id}'s question: {question}")
        logger.info(f"the user_{user_id}'s correct answer: {answer}")

        upload_last_question(user_id, question, r_conn)
        upload_last_correct_answer(user_id, answer, r_conn)

        # удаляем выданный вопрос
        delete_select_question(user_id, num_qa, r_conn)
        delete_select_correct_answer(user_id, num_qa, r_conn)
        return 1
    return 0

def handle_successfully_uploaded_question(vk, user_id, r_conn):
    question = get_last_question(user_id, r_conn)
    vk.messages.send(user_id=user_id,
                     message=question,
                     random_id=get_random_id(),
                     keyboard=create_keyboard())

def handle_incorrect_upload_question(vk, user_id):
    message = "Прошли все вопросы, приходите позднее.."
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id())

def func_user_answer(vk, SBER_TOKEN, received_message, user_id, r_conn):
    current_time = datetime.datetime.now()
    response_time = current_time.strftime("%Y-%m-%d %H:%M:%S.%f")

    # print("Дата и время ответа:", response_time)

    question = get_last_question(user_id, r_conn)
    correct_answer = get_last_correct_answer(user_id, r_conn)
    correct_answer = scale_text(correct_answer).lower().strip('.')

    # print('Вопрос:', question)
    # print('Правильный ответ:', correct_answer)
    # print('Пользовательский ответ:', received_message)

    logger.info(f"the user_{user_id}'s answer: {received_message}")
    cos_sim = get_cosine_similarity(received_message, correct_answer, SBER_TOKEN)
    # print(cos_sim)
    # similarity_score = get_similarity_score(received_message, correct_answer, question, llm)
    similarity_score, reasoning = custom_evaluate_qa(received_message, correct_answer, question, SBER_TOKEN)

    # print(cos_sim, similarity_score, reasoning)
    if cos_sim >= 0.98 or similarity_score > 0.8:
        bonus = 100
        upload_account(bonus, user_id, r_conn)
        message = f"Ответили верно! Получате {bonus} балл(а/ов).\n Причина: {reasoning}"
        vk.messages.send(user_id=user_id,
                         message=message,
                         random_id=get_random_id(),
                         keyboard=create_keyboard())

    elif cos_sim >= 0.97 or similarity_score >= 0.7:
        bonus = 70
        upload_account(bonus, user_id, r_conn)
        message = f"Ответили, почти верно! Получате {bonus} балл(а/ов).\n Причина: {reasoning}"
        vk.messages.send(user_id=user_id,
                         message=message,
                         random_id=get_random_id(),
                         keyboard=create_keyboard())

    elif similarity_score >= 0.5:
        bonus = 50
        upload_account(bonus, user_id, r_conn)
        message = f"Ответили, частично верно! Получате {bonus} балл(а/ов).\n Причина: {reasoning}"
        vk.messages.send(user_id=user_id,
                         message=message,
                         random_id=get_random_id(),
                         keyboard=create_keyboard())

    elif similarity_score >= 0.3:
        bonus = 30
        upload_account(bonus, user_id, r_conn)
        message = f"Ответили, почти неверно! Получате {bonus} балл(а/ов).\n Причина: {reasoning}"
        vk.messages.send(user_id=user_id,
                         message=message,
                         random_id=get_random_id(),
                         keyboard=create_keyboard())

    elif similarity_score >= 0.1:
        bonus = 10
        upload_account(bonus, user_id, r_conn)
        message = f"Ммм... Получате {bonus} балл(а/ов).\n Причина: {reasoning}"
        vk.messages.send(user_id=user_id,
                         message=message,
                         random_id=get_random_id(),
                         keyboard=create_keyboard())

    else:
        message = f"Нет, неверно! \n Причина: {reasoning}"
        vk.messages.send(user_id=user_id,
                         message=message,
                         random_id=get_random_id(),
                         keyboard=create_keyboard())

    # сохранение ответов пользователя
    qa_info = json.dumps({
        'question': question,
        'answer': received_message
    })

    r_conn.hset(
        f'answers_id_{user_id}',
        response_time,
        qa_info
    )

    # correct_answer_num = r_conn.hget('num_of_last_question', f'id_{user_id}')
    # db.delete_select_correct_answer(user_id, correct_answer_num, r_conn)

def func_show_user_account(vk, user_id, r_conn):
    users_info_dict = json.loads(
        r_conn.hget('users_info', f'id_{user_id}')
    )

    total_account = int(users_info_dict['account'])

    if total_account is not None:
        message = f"На счете у вас {total_account} балл(а/ов)."
        vk.messages.send(user_id=user_id,
                         message=message,
                         random_id=get_random_id(),
                         keyboard=create_keyboard())
    else:
        message = "Пока у вас нет бонусных баллов."
        vk.messages.send(user_id=user_id,
                         message=message,
                         random_id=get_random_id(),
                         keyboard=create_keyboard())

def switch_ruzik_chat(user_id, r_conn):
    if r_conn.hget('ruzik_chat_keys', f'id_{user_id}') == 'on':
        return True
    return False

def switch_admin(user_id, r_conn):
    if r_conn.hget(f'admin_id_{user_id}', 'admin_key') == 'on':
        return True
    return False

def handle_ruzik_chat(vk, SBER_TOKEN, received_message, user_id, r_conn):
    messages = []
    if r_conn.hgetall('ruzik_chat_messages') and r_conn.hget('ruzik_chat_messages', f'id_{user_id}') is not None:
        messages = json.loads(r_conn.hget('ruzik_chat_messages', f'id_{user_id}')).get('messages')
        # messages = r_conn.hget('ruzik_chat_messages', f'id_{user_id}')

    chat_token = get_token(SBER_TOKEN)

    response, messages = connect_ruzik_chat(chat_token, received_message, messages)
    # r_conn.hset('ruzik_chat_messages', f'id_{user_id}', messages)
    json_messages = json.dumps({'messages': messages})

    r_conn.hset('ruzik_chat_messages', f'id_{user_id}', json_messages)

    message = messages[-1]['content']
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id())

def handle_admin_login(vk, user_id):
    message = "Введите логин:"
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id())

def handle_incorrect_admin_login(vk, user_id):
    message = "Неправильно введен логин. Попробуйте еще раз. Если хотите выйти из администратора, напишите слово 'выйти'."
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id())

def handle_incorrect_admin_password(vk, user_id):
    message = "Неправильно введен пароль. Попробуйте еще раз. Если хотите выйти из администратора, напишите слово 'выйти'."
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id())

def handle_incorrect_func(vk, user_id):
    message = "Неправильно управляйте функциями. Попробуйте еще раз."
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_qa_admin_keyboard())

def handle_admin_password(vk, user_id):
    message = "Введите пароль:"
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id())

def handle_admin_start(vk, user_id):
    message = "Вошли в админстратор."
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_admin_keyboard())

def handle_admin_stop(vk, user_id):
    message = "Вышли из админстратора."
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id())

def handle_upload_qa(vk, user_id):
    message = ('Загрузите файл в формате .txt .'
               '\n Каждый вопрос должен начинаться со строки "Вопрос: ", а ответ - со строки "Ответ: ". Все вопрос-ответы должны разделены пустыми строками.'
               '\n\n Пример:'
               '\n Вопрос: В каком году была основана наша компания?'
               '\n Ответ: Наша компания была основана в 1995 году.'
               '\n\n Вопрос: Какое название бренда у наших кукурузных палочек?'
               '\nОтвет: Название нашего бренда - "Золотые Палочки".')
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_qa_admin_keyboard())

def handle_successfully_uploaded_qa(vk, user_id):
    message = f"Файл успешно загружен."
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_admin_keyboard())

def handle_successfully_deleted_qa(vk, user_id):
    message = f"Все вопросы и ответы успешно удалены."
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_admin_keyboard())
def handle_no_deleted_qa(vk, user_id):
    message = f"Вопросы и ответы не удалены."
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_admin_keyboard())

def handle_request_delete_qa(vk, user_id):
    message = "Действительно ли хотите удалить все файлы?"
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_request_admin_keyboard())

def handle_request_edit_qa(vk, user_id):
    message = "Введите id пользователя, которого хотите отредактировать в вопросах и ответах (id должен состоять только из цифр)."
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_qa_admin_keyboard())

def handle_change_login(vk, user_id):
    message = "Введите новый логин:"
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id())

def handle_successfully_changed_login(vk, received_message, user_id):
    message = f"Успешно изменен логин администратора. \nВаш новый логин: {received_message}"
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id())

def handle_incorrect_changed_login(vk, user_id):
    message = f"Не можете менять логин на слово 'админ'. Введите еще раз новый логин."
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id())

def handle_change_password(vk, user_id):
    message = "Введите новый пароль:"
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id())

def handle_successfully_changed_password(vk, received_message, user_id):
    message = f"Успешно изменен пароль администратора. \nВаш новый пароль: {received_message}"
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_admin_keyboard())

def handle_func_qa(vk, user_id):
    message = "Выберите нужную функцию."
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_qa_admin_keyboard())
def handle_func_generate_qa(vk, user_id):
    message = "Загрузите txt файл."
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_admin_keyboard())

def handle_func_num_qa(vk, user_id):
    message = "Введите количество вопросов, которые вы хотите получить."
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id())

def handle_back_admin(vk, user_id):
    message = "Вернулись в главный экран."
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_admin_keyboard())
def handle_successfully_get_users_info(vk, peer_id, title):
    upload_url = vk.docs.getMessagesUploadServer(type='doc', peer_id=peer_id)['upload_url']
    req = requests.post(upload_url, files={'file': open(title, 'rb')}).json()
    file = vk.docs.save(file=req['file'])

    vk.messages.send(peer_id=peer_id,
                     attachment=f"doc{file['doc']['owner_id']}_{file['doc']['id']}",
                     random_id=get_random_id(),
                     keyboard=create_admin_keyboard())

def handle_successfully_get_generated_qa(vk, peer_id, text_file):
    title = "example_qa.txt"

    with open(title, "w", encoding="utf-8") as file:
        file.write(text_file)

    upload_url = vk.docs.getMessagesUploadServer(type='doc', peer_id=peer_id)['upload_url']
    # req = requests.post(upload_url, files={'file': open(title, 'rb')}).json()
    req = requests.post(upload_url, files={'file': open(title, 'rb')}).json()
    file = vk.docs.save(file=req['file'])

    vk.messages.send(peer_id=peer_id,
                     attachment=f"doc{file['doc']['owner_id']}_{file['doc']['id']}",
                     random_id=get_random_id(),
                     keyboard=create_admin_keyboard())

def handle_successfully_uploaded_users_info(vk, user_id):
    message = f"Файл успешно загружен."
    vk.messages.send(user_id=user_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_admin_keyboard())

def handle_request_upload_users_info(vk, admin_id):
    message = "Загрузив xlsx файл, можете заменить информации о данных пользователей."
    vk.messages.send(user_id=admin_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_admin_keyboard())

def exit_admin(user_id, r_conn):
    r_conn.hset(f'admin_id_{user_id}', 'login_key', 'off')
    r_conn.hset(f'admin_id_{user_id}', 'password_key', 'off')
    r_conn.hset(f'admin_id_{user_id}', 'func_upload_qa_key', 'off')
    r_conn.hset(f'admin_id_{user_id}', 'request_delete_key', 'off')
    r_conn.hset(f'admin_id_{user_id}', 'func_delete_qa_key', 'off')
    r_conn.hset(f'admin_id_{user_id}', 'func_edit_key', 'off')
    r_conn.hset(f'admin_id_{user_id}', 'func_qa_key', 'off')
    r_conn.hset(f'admin_id_{user_id}', 'admin_key', 'off')
    r_conn.hset(f'admin_id_{user_id}', 'users_info_key', 'off')
    r_conn.hset(f'admin_id_{user_id}', 'generate_qa_key', 'off')
    r_conn.hset(f'admin_id_{user_id}', 'num_qa_key', 'off')
    r_conn.hset(f'admin_id_{user_id}', 'change_login_key', 'off')
    r_conn.hset(f'admin_id_{user_id}', 'change_password_key', 'off')
    r_conn.hdel(f'admin_id_{user_id}', 'edit_user_id')

def check_login_key(user_id, r_conn):
    if r_conn.hget(f'admin_id_{user_id}', 'login_key') == 'on':
        return True
    return False

def check_password_key(user_id, r_conn):
    if r_conn.hget(f'admin_id_{user_id}', 'password_key') == 'on':
        return True
    return False


def check_qa(user_id, r_conn):
    if r_conn.hget(f'admin_id_{user_id}', 'func_qa_key') == 'on':
        return True
    return False

def check_func_generate_qa(user_id, r_conn):
    if r_conn.hget(f'admin_id_{user_id}', 'generate_qa_key') == 'on':
        return True
    return False

def check_num_qa(user_id, r_conn):
    if r_conn.hget(f'admin_id_{user_id}', 'num_qa_key') == 'on':
        return True
    return False

def check_upload_key(user_id, r_conn):
    if r_conn.hget(f'admin_id_{user_id}', 'func_upload_qa_key') == 'on':
        return True
    return False

def check_users_info(user_id, r_conn):
    if r_conn.hget(f'admin_id_{user_id}', 'users_info_key') == 'on':
        return True
    return False
def check_clear_qa(user_id, r_conn):
    if r_conn.hget(f'admin_id_{user_id}', 'func_delete_qa_key') == 'on':
        return True
    return False

def check_edit_qa(user_id, r_conn):
    if r_conn.hget(f'admin_id_{user_id}', 'func_edit_key') == 'on':
        return True
    return False
def check_admin_login(received_message, admin_id, r_conn):
    if r_conn.hget(f'admin_id_{admin_id}', 'login') == received_message:
        return True
    return False

def check_admin_password(received_message, admin_id, r_conn):
    if r_conn.hget(f'admin_id_{admin_id}', 'password') == received_message:
        return True
    return False

def check_admin_change_login(admin_id, r_conn):
    if r_conn.hget(f'admin_id_{admin_id}', 'change_login_key') == 'on':
        return True
    return False

def check_admin_change_password(admin_id, r_conn):
    if r_conn.hget(f'admin_id_{admin_id}', 'change_password_key') == 'on':
        return True
    return False

def str_to_dict(value):
    return json.loads(value)

def handle_successfully_get_qa(vk, admin_id, title):
    upload_url = vk.docs.getMessagesUploadServer(type='doc', peer_id=admin_id)['upload_url']
    req = requests.post(upload_url, files={'file': open(title, 'rb')}).json()
    file = vk.docs.save(file=req['file'])

    vk.messages.send(peer_id=admin_id,
                     attachment=f"doc{file['doc']['owner_id']}_{file['doc']['id']}",
                     random_id=get_random_id(),
                     keyboard=create_qa_admin_keyboard())

def handle_request_upload_qa(vk, admin_id):
    message = "Загрузив xlsx файл с таким же форматом, можете заменить набор вопросов и ответов для текущего пользователя."
    vk.messages.send(user_id=admin_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_qa_admin_keyboard())
def handle_incorrect_edit_qa(vk, admin_id):
    message = 'Неправильно ввели id пользователя. Введите еще раз корректный id пользователя.'
    vk.messages.send(user_id=admin_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_qa_admin_keyboard())

def handle_incorrect_users_info(vk, admin_id):
    message = 'Файл был загружен неправильно, файл должен быть в формате xlsx. Попробуйте еще раз.'
    vk.messages.send(user_id=admin_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_qa_admin_keyboard())

def handle_incorrect_num_qa(vk, admin_id):
    message = 'Число введено неверно. Попробуйте еще раз.'
    vk.messages.send(user_id=admin_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_admin_keyboard())
def handle_successfully_back(vk, admin_id):
    message = 'Выберите нужную функцию.'
    vk.messages.send(user_id=admin_id,
                     message=message,
                     random_id=get_random_id(),
                     keyboard=create_admin_keyboard())
def admin_func_upload_qa(event, r_conn):
    event_obj = event.obj['message']
    document = event_obj['attachments']
    if len(document) >= 1 and document[0]['type'] == 'doc':
        title = document[0]['doc']['title']
        url = document[0]['doc']['url']
        directory_name = 'questions_data'
        path_directory = os.path.abspath(f'./{directory_name}')
        path_directory = path_directory.replace('\\', '/')
        urllib.request.urlretrieve(url, f'{path_directory}/{title}')
        users_ids = list(r_conn.hgetall('users_info').keys())
        upload_one_file_of_qa(users_ids, r_conn, title)
        return 1
    return 0

def admin_func_delete_qa(vk, event, r_conn):
    event_obj = event.obj['message']
    user_id = event_obj['from_id']
    received_message = event_obj['text']

    if received_message == "Да":
        users_id = list(r_conn.hgetall('users_info').keys())
        clear_questions(users_id, r_conn)
        clear_answers(users_id, r_conn)
        clear_qa_from_dir()
        r_conn.hset(f'admin_id_{user_id}', 'func_delete_qa_key', 'off')
        handle_successfully_deleted_qa(vk, user_id)
    else:
        r_conn.hset(f'admin_id_{user_id}', 'func_delete_qa_key', 'off')
        handle_no_deleted_qa(vk, user_id)
def upload_xlsx_file_of_qa(user_id, r_conn, title):
    directory_name = 'questions_data'
    path_directory = os.path.abspath(f'./{directory_name}')
    path_directory = path_directory.replace('\\', '/')

    workbook = openpyxl.load_workbook(f'{path_directory}/{title}')
    worksheet = workbook.active
    qa_counter = 0

    # Загрузка данных в Redis
    for row in range(2, worksheet.max_row + 1):
        question = worksheet.cell(row=row, column=1).value
        answer = worksheet.cell(row=row, column=2).value
        r_conn.hset(f'questions_{user_id}', f'{qa_counter}', question)
        r_conn.hset(f'correct_answers_{user_id}', f'{qa_counter}', answer)
        qa_counter += 1

def upload_xlsx_file_of_users_info(r_conn, title):
    directory_name = 'questions_data'
    path_directory = os.path.abspath(f'./{directory_name}')
    path_directory = path_directory.replace('\\', '/')

    workbook = openpyxl.load_workbook(f'{path_directory}/{title}')
    worksheet = workbook.active

    # Загрузка данных в Redis
    for row in range(2, worksheet.max_row + 1):
        user_id = worksheet.cell(row=row, column=2).value
        first_name = worksheet.cell(row=row, column=3).value
        last_name = worksheet.cell(row=row, column=4).value
        b_date = worksheet.cell(row=row, column=5).value
        sex = worksheet.cell(row=row, column=6).value
        city = worksheet.cell(row=row, column=7).value
        country = worksheet.cell(row=row, column=8).value
        account = worksheet.cell(row=row, column=9).value
        info_dict = {'first_name': first_name,
                     'last_name': last_name,
                     'b_date': b_date,
                     'sex': sex,
                     'city': city,
                     'country': country,
                     'account': account
                     }

        r_conn.hset('users_info',
                    user_id,
                    json.dumps(info_dict))
def admin_func_edit_qa(event, r_conn):
    event_obj = event.obj['message']
    document = event_obj['attachments']
    admin_id = event_obj['from_id']
    if len(document) >= 1 and document[0]['type'] == 'doc':
        if r_conn.hgetall(f'admin_id_{admin_id}') and r_conn.hget(f'admin_id_{admin_id}', 'edit_user_id') is not None:
            user_id = r_conn.hget(f'admin_id_{admin_id}', 'edit_user_id')
            title = document[0]['doc']['title']
            url = document[0]['doc']['url']
            directory_name = 'questions_data'
            path_directory = os.path.abspath(f'./{directory_name}')
            path_directory = path_directory.replace('\\', '/')
            urllib.request.urlretrieve(url, f'{path_directory}/{title}')
            r_conn.delete(f'questions_{user_id}')
            r_conn.delete(f'correct_answers_{user_id}')
            r_conn.hdel(f'admin_id_{admin_id}', 'edit_user_id')
            upload_xlsx_file_of_qa(user_id, r_conn, title)
            os.remove(f'{path_directory}/{title}')
            return 2, None
        else:
            return 0, None

    elif event.type == VkBotEventType.MESSAGE_NEW:
        event_obj = event.obj['message']
        admin_id = event_obj['from_id']
        user_id = event_obj['text']
        user_id = f'id_{user_id}'
        r_conn.hset(f'admin_id_{admin_id}', 'edit_user_id', user_id)

        if user_id not in list(r_conn.hgetall('users_info').keys()):
            return 0, None

        questions, answers = get_user_qa(user_id, r_conn)
        questions_df = pd.DataFrame.from_dict({**questions}, orient='index',  columns=['questions'])
        answers_df = pd.DataFrame.from_dict({**answers}, orient='index',  columns=['answers'])
        qa_df = pd.concat([questions_df, answers_df], names=['questions', 'answers'], axis=1)

        title = f'qa_{user_id}.xlsx'
        qa_df.to_excel(title, index=False)
        return 1, title
    return 0, None

def admin_func_upload_users_info(event, r_conn):
    event_obj = event.obj['message']
    document = event_obj['attachments']
    if len(document) >= 1 and document[0]['type'] == 'doc':
        if r_conn.hgetall('users_info'):
            title = document[0]['doc']['title']
            url = document[0]['doc']['url']
            directory_name = 'questions_data'
            path_directory = os.path.abspath(f'./{directory_name}')
            path_directory = path_directory.replace('\\', '/')
            urllib.request.urlretrieve(url, f'{path_directory}/{title}')
            r_conn.delete('users_info')
            upload_xlsx_file_of_users_info(r_conn, title)
            return 1
    return 0

def admin_func_get_users_info(r_conn):
    users_info_dict = get_users_info(r_conn)
    df = pd.DataFrame.from_dict(users_info_dict, orient='index').reset_index()
    df.columns = ['id', 0]
    df_id = df['id']
    df = df[0].apply(str_to_dict).apply(pd.Series)
    df = pd.concat([df_id, df], axis=1)
    df.columns = ['id', 'first_name', 'last_name', 'b_date', 'sex', 'city', 'country', 'account']
    df.to_excel('users_info_file.xlsx')

def admin_func_qa(vk, event, r_conn):
    event_obj = event.obj['message']
    user_id = event_obj['from_id']
    received_message = event_obj['text']
    if received_message == "Добавлять вопросы и ответы":
        r_conn.hset(f'admin_id_{user_id}', 'func_upload_qa_key', 'on')
        r_conn.hset(f'admin_id_{user_id}', 'func_edit_key', 'off')
        r_conn.hset(f'admin_id_{user_id}', 'func_delete_qa_key', 'off')
        handle_upload_qa(vk, user_id)
    elif received_message == "Редактировать вопросы и ответы":
        r_conn.hset(f'admin_id_{user_id}', 'func_edit_key', 'on')
        r_conn.hset(f'admin_id_{user_id}', 'func_upload_qa_key', 'off')
        r_conn.hset(f'admin_id_{user_id}', 'func_delete_qa_key', 'off')
        handle_request_edit_qa(vk, user_id)
    elif received_message == "Удалять вопросы и ответы":
        r_conn.hset(f'admin_id_{user_id}', 'func_delete_qa_key', 'on')
        r_conn.hset(f'admin_id_{user_id}', 'func_upload_qa_key', 'off')
        r_conn.hset(f'admin_id_{user_id}', 'func_edit_key', 'off')
        handle_request_delete_qa(vk, user_id)
    elif received_message == "Назад":
        r_conn.hset(f'admin_id_{user_id}', 'func_qa_key', 'off')
        r_conn.hset(f'admin_id_{user_id}', 'func_upload_qa_key', 'off')
        r_conn.hset(f'admin_id_{user_id}', 'func_edit_key', 'off')
        r_conn.hset(f'admin_id_{user_id}', 'func_delete_qa_key', 'off')
        r_conn.hdel(f'admin_id_{user_id}', 'edit_user_id')
        handle_successfully_back(vk, user_id)
    elif check_upload_key(user_id, r_conn):
        admin_func_upload_qa(event, r_conn)
        r_conn.hset(f'admin_id_{user_id}', 'func_upload_qa_key', 'off')
        handle_successfully_uploaded_qa(vk, user_id)
    elif check_clear_qa(user_id, r_conn):
        admin_func_delete_qa(vk, event, r_conn)
        r_conn.hset(f'admin_id_{user_id}', 'func_qa_key', 'off')
    elif check_edit_qa(user_id, r_conn):
        event_obj = event.obj['message']
        admin_id = event_obj['peer_id']
        func, title = admin_func_edit_qa(event, r_conn)
        if func == 1:
            handle_successfully_get_qa(vk, admin_id, title)
            handle_request_upload_qa(vk, admin_id)
            os.remove(title)
        elif func == 2:
            handle_successfully_uploaded_qa(vk, admin_id)
            r_conn.hset(f'admin_id_{user_id}', 'func_edit_key', 'off')
        else:
            handle_incorrect_edit_qa(vk, admin_id)
    else:
        handle_incorrect_func(vk, user_id)
def admin_func_upload_text(event, r_conn):
    event_obj = event.obj['message']
    document = event_obj['attachments']
    admin_id = event_obj['from_id']
    if len(document) >= 1 and document[0]['type'] == 'doc':
        title = document[0]['doc']['title']
        url = document[0]['doc']['url']
        urllib.request.urlretrieve(url, title)
        r_conn.hset(f'admin_id_{admin_id}', 'text_name', title)
        return 1
    return 0

def main():

    load_dotenv()
    VK_USER_TOKEN = os.getenv('VK_USER_TOKEN')
    VK_GROUP_TOKEN = os.getenv('VK_GROUP_TOKEN')
    GROUP_ID = os.getenv('GROUP_ID')
    SBER_TOKEN = os.getenv('SBER_TOKEN')
    host = os.getenv('host')
    port = os.getenv('port')
    db = 0

    r_conn = redis.Redis(
        host=host,
        port=port,
        db=db,
        charset='utf-8',
        decode_responses=True
    )

    logging.basicConfig(
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        level=logging.INFO,
        filename='py_log.log',
        filemode='w'
    )

    while True:
        try:
            logger.debug('the bot started')
            authorize = vk_api.VkApi(token=VK_GROUP_TOKEN)
            vk = authorize.get_api()
            longpoll = VkBotLongPoll(authorize, group_id=GROUP_ID)
            for event in longpoll.listen():
                # новые текстовые сообщения, адресованные ему
                if event.type == VkBotEventType.MESSAGE_NEW:
                    event_obj = event.obj['message']
                    received_message = event_obj['text'] # это и есть ответ
                    # уникальный идентификационный номер
                    user_id = event_obj['from_id'] # id участника
                    peer_id = event_obj['peer_id']

                    if not r_conn.hgetall('ruzik_chat_keys') or r_conn.hget('ruzik_chat_keys', f'id_{user_id}') is None:
                        r_conn.hset('ruzik_chat_keys', f'id_{user_id}', 'on')

                    if received_message in ['админ', 'вход в админ', 'Админ']:
                        if check_admin_change_login(user_id, r_conn):
                            handle_incorrect_changed_login(vk, user_id)
                            continue

                        if not r_conn.hgetall(f'admin_id_{user_id}'):
                            r_conn.hset(f'admin_id_{user_id}', 'login', 'ruzik_admin')
                            r_conn.hset(f'admin_id_{user_id}', 'password', '0000')
                            r_conn.hset(f'admin_id_{user_id}', 'change_login_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'change_password_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'login_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'password_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'func_upload_qa_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'request_delete_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'func_delete_qa_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'admin_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'func_qa_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'users_info_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'generate_qa_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'num_qa_key', 'off')

                        r_conn.hset(f'admin_id_{user_id}', 'login_key', 'on')
                        handle_admin_login(vk, user_id)
                    elif ((check_login_key(user_id, r_conn) or check_password_key(user_id, r_conn) or switch_admin(user_id, r_conn))
                          and received_message in ["выйти", "Выйти"]):
                        exit_admin(user_id, r_conn)
                        handle_admin_stop(vk, user_id)
                    elif check_login_key(user_id, r_conn):
                        if check_admin_login(received_message, user_id, r_conn):
                            r_conn.hset(f'admin_id_{user_id}', 'login_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'password_key', 'on')
                            handle_admin_password(vk, user_id)
                        else:
                            handle_incorrect_admin_login(vk, user_id)
                    elif check_password_key(user_id, r_conn):
                        if check_admin_password(received_message, user_id, r_conn):
                            r_conn.hset(f'admin_id_{user_id}', 'password_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'admin_key', 'on')
                            handle_admin_start(vk, user_id)
                        else:
                            handle_incorrect_admin_password(vk, user_id)
                    elif switch_admin(user_id, r_conn):
                        if received_message == "Данные игроков":
                            admin_func_get_users_info(r_conn)
                            r_conn.hset(f'admin_id_{user_id}', 'users_info_key', 'on')
                            r_conn.hset(f'admin_id_{user_id}', 'func_qa_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'generate_qa_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'change_login_key', 'off')
                            title = 'users_info_file.xlsx'
                            handle_successfully_get_users_info(vk, peer_id, title)
                            handle_request_upload_users_info(vk, user_id)
                            # os.remove(title)
                        elif received_message == "Вопросы и ответы":
                            r_conn.hset(f'admin_id_{user_id}', 'func_qa_key', 'on')
                            r_conn.hset(f'admin_id_{user_id}', 'users_info_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'generate_qa_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'num_qa_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'change_login_key', 'off')
                            handle_func_qa(vk, user_id)
                        elif received_message == "Генерация вопросов":
                            r_conn.hset(f'admin_id_{user_id}', 'generate_qa_key', 'on')
                            r_conn.hset(f'admin_id_{user_id}', 'num_qa_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'func_qa_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'users_info_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'change_login_key', 'off')
                            handle_func_generate_qa(vk, user_id)
                        elif received_message == "Изменить логин и пароль":
                            r_conn.hset(f'admin_id_{user_id}', 'change_login_key', 'on')
                            r_conn.hset(f'admin_id_{user_id}', 'func_qa_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'users_info_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'generate_qa_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'num_qa_key', 'off')
                            handle_change_login(vk, user_id)
                        elif check_admin_change_login(user_id, r_conn):
                            r_conn.hset(f'admin_id_{user_id}', 'login', received_message)
                            r_conn.hset(f'admin_id_{user_id}', 'change_login_key', 'off')
                            r_conn.hset(f'admin_id_{user_id}', 'change_password_key', 'on')
                            handle_successfully_changed_login(vk, received_message, user_id)
                            handle_change_password(vk, user_id)
                        elif check_admin_change_password(user_id, r_conn):
                            r_conn.hset(f'admin_id_{user_id}', 'password', received_message)
                            r_conn.hset(f'admin_id_{user_id}', 'change_password_key', 'off')
                            handle_successfully_changed_password(vk, received_message, user_id)
                        elif check_users_info(user_id, r_conn):
                            if admin_func_upload_users_info(event, r_conn):
                                handle_successfully_uploaded_users_info(vk, user_id)
                            else:
                                handle_incorrect_users_info(vk, user_id)
                        elif check_qa(user_id, r_conn):
                            admin_func_qa(vk, event, r_conn)
                        elif check_func_generate_qa(user_id, r_conn):
                            admin_func_upload_text(event, r_conn)
                            handle_func_num_qa(vk, user_id)
                            r_conn.hset(f'admin_id_{user_id}', 'num_qa_key', 'on')
                            r_conn.hset(f'admin_id_{user_id}', 'generate_qa_key', 'off')
                        elif check_num_qa(user_id, r_conn):
                            if received_message.isdigit():
                                num = received_message
                                title = r_conn.hget(f'admin_id_{user_id}', 'text_name')
                                loader = TextLoader(title, encoding='UTF-8')
                                doc = loader.load()
                                text_file = custom_generate_qa(doc, int(num), SBER_TOKEN)
                                handle_successfully_get_generated_qa(vk, user_id, text_file)
                                handle_func_num_qa(vk, user_id)
                            elif received_message in ["назад", "Назад"]:
                                handle_back_admin(vk, user_id)
                                r_conn.hset(f'admin_id_{user_id}', 'num_qa_key', 'off')
                                # os.remove(title)
                            else:
                                handle_incorrect_num_qa(vk, user_id)
                        else:
                            handle_successfully_back(vk, user_id)

                    elif received_message in ["начать", "start", "старт", "Начать", "Start", "Старт"]:
                        user_func_start(VK_USER_TOKEN, user_id, r_conn)
                        handle_user_start(vk, user_id)
                        # off the ruzik chat
                        r_conn.hset('ruzik_chat_keys', f'id_{user_id}', 'off')
                    elif received_message in ["закончить", "стоп", "end", "stop", "Закончить", "Стоп", "End", "Stop"]:
                        handle_user_stop(vk, user_id)
                        # on the ruzik chat
                        r_conn.hset('ruzik_chat_keys', f'id_{user_id}', 'on')
                    elif switch_ruzik_chat(user_id, r_conn):
                        handle_ruzik_chat(vk, SBER_TOKEN, received_message, user_id, r_conn)
                    elif received_message == "Вопрос":
                        if func_question(user_id, r_conn):
                            handle_successfully_uploaded_question(vk, user_id, r_conn)
                        else:
                            handle_incorrect_upload_question(vk, user_id)
                            # on the ruzik chat
                            r_conn.hset('ruzik_chat_keys', f'id_{user_id}', 'on')
                    elif received_message == "На счете":
                        func_show_user_account(vk, user_id, r_conn)
                    else:
                        func_user_answer(vk, SBER_TOKEN, received_message, user_id, r_conn)

        except KeyboardInterrupt:
            logger.info('the bot stopped.')

        except Exception as exception:
            logger.error('the bot crashed with an error.')
            logger.exception(exception)

if __name__ == "__main__":
    main()